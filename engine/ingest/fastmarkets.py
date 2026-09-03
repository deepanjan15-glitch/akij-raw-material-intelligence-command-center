"""Ingest Fastmarkets price assessment sheets -> normalized PriceObservation entities."""
from __future__ import annotations
import openpyxl
from datetime import datetime, date
from ..core.entities import (
    PriceObservation, extract_country, extract_incoterm, extract_unit, extract_specification,
)

FM_FILE = "/Users/deep/Downloads/Raw_Material.xlsx"

# Sheet name -> (family material name, category)
SHEET_FAMILY = {
    "Corn_Or_Maize": ("Corn/Maize", "Grains & Cereals"),
    "Corn_Gluten_Feed _CGF": ("Corn Gluten Feed", "Grains & Cereals"),
    "Corn_Gluten_Meal_or_CGM": ("Corn Gluten Meal", "Grains & Cereals"),
    "Distiller's_Corn_Oil": ("Distiller's Corn Oil", "Oils & Fats"),
    "Corn_DDGS": ("Corn DDGS", "Grains & Cereals"),
    "Rapeseed_Or_Canola": ("Rapeseed/Canola", "Oilseeds & Meals"),
    "Rapeseed_Or_Canola_Meal": ("Rapeseed Meal", "Oilseeds & Meals"),
    "Rapeseed_Or_Canola_Oil": ("Rapeseed Oil", "Oils & Fats"),
    "Barley_Or_Feed_barley": ("Barley", "Grains & Cereals"),
    "Cotton_Seed": ("Cotton Seed", "Oilseeds & Meals"),
    "Sunflower_Meal": ("Sunflower Meal", "Oilseeds & Meals"),
    "Sunflower_Oil": ("Sunflower Oil", "Oils & Fats"),
    "Wheat": ("Wheat", "Grains & Cereals"),
    "Palm_Oil": ("Palm Oil", "Oils & Fats"),
    "Soymeal_Or_SBM": ("Soymeal", "Oilseeds & Meals"),
    "Soybean": ("Soybean", "Oilseeds & Meals"),
    "Soybean_Oil": ("Soybean Oil", "Oils & Fats"),
    "Fish_Oil": ("Fish Oil", "Oils & Fats"),
}

# USD/MT conversion factors (documented)
SHORT_TON_TO_MT = 1.10231131
CTS_LB_TO_MT = 22.0462262
BUSHEL = {"CRN": 39.3683, "WHE": 36.7437, "SYB": 36.7437, "RPS": 36.7437, "RSD": 36.7437,
          "SSD": 36.7437, "CTS": 36.7437, "BRY": 36.7437, "PLM": 36.7437}


def _num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _to_usd_mt(symbol: str, desc: str, value) -> float | None:
    if value is None:
        return None
    d = desc.lower()
    prefix = symbol.split("-")[1] if "-" in symbol else ""
    if "$/short ton" in d:
        return value * SHORT_TON_TO_MT
    if "cts/lb" in d or "cents/lb" in d or "c/lb" in d:
        return value * CTS_LB_TO_MT
    if "$/bushel" in d:
        return value * BUSHEL.get(prefix, 36.7437)
    if "c$/bu" in d or "c/bu" in d:
        return value / 100.0 * BUSHEL.get(prefix, 36.7437)
    if "€" in d or "euro/mt" in d or "hryvnia" in d or "real/" in d or "ringgit" in d or "rupiah" in d:
        return None  # not USD-convertible without FX; keep original only
    return value  # $/mt or $/tonne


def ingest_fastmarkets(loaded_at: str) -> list[PriceObservation]:
    wb = openpyxl.load_workbook(FM_FILE, data_only=True)
    out: list[PriceObservation] = []
    n = 0
    for sheet, (family, category) in SHEET_FAMILY.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        # locate header + column indices
        header_idx, col = None, {}
        for ri, row in enumerate(rows):
            for ci, v in enumerate(row):
                if v == "Symbol":
                    header_idx = ri
                    break
            if header_idx is not None:
                break
        if header_idx is None:
            continue
        for ci, v in enumerate(rows[header_idx]):
            if v == "Symbol":
                col["symbol"] = ci
            elif v == "Description":
                col["desc"] = ci
            elif v == "Mid":
                col["mid"] = ci
            elif v == "Monthly Average":
                col["mavg"] = ci
            elif v in ("Assessment Date", "Date"):
                col["date"] = ci
        if "symbol" not in col or "mid" not in col:
            continue
        for row in rows[header_idx + 1:]:
            sym = row[col["symbol"]] if len(row) > col["symbol"] else None
            if not isinstance(sym, str) or not sym.startswith("AG-"):
                continue
            desc = str(row[col["desc"]] or "") if "desc" in col and len(row) > col["desc"] else ""
            mid = _num(row[col["mid"]]) if "mid" in col and len(row) > col["mid"] else None
            mavg = _num(row[col["mavg"]]) if "mavg" in col and len(row) > col["mavg"] else None
            if mid is None and mavg is None:
                continue
            raw_date = row[col["date"]] if "date" in col and len(row) > col["date"] else None
            if isinstance(raw_date, datetime):
                obs_date = raw_date.date().isoformat()
            elif isinstance(raw_date, date):
                obs_date = raw_date.isoformat()
            else:
                obs_date = ""
            country_code, country_name, market = extract_country(desc)
            incoterm = extract_incoterm(desc)
            currency, unit = extract_unit(desc)
            spec = extract_specification(desc)
            for tag, val in (("current", mid), ("monthly-avg", mavg)):
                if val is None:
                    continue
                n += 1
                out.append(PriceObservation(
                    id=f"FM-{sym}-{tag}",
                    materialId=None,          # mapped later against the 87-master
                    sourceId="fastmarkets",
                    countryCode=country_code,
                    countryName=country_name,
                    incoterm=incoterm,
                    currency=currency,
                    unit=unit,
                    value=val,
                    valueUsdMt=_to_usd_mt(sym, desc, val),
                    market=market,
                    specification=spec or (desc if desc else None),
                    observationDate=obs_date,
                    loadedAt=loaded_at,
                ))
                out[-1]._family = family
    return out
