"""Ingest the dated price-dashboard snapshots -> per-material price history.

Source: "Feed Raw Material Price Dashboard..xlsx" — one dated sheet per snapshot
(each sheet is a flat Input File with the same "Product / Current Avg Price"
layout, though the oldest sheets lack the "Source" column -> schema drift).

Only exact (case-insensitive) product-name matches are kept; names are never
forced/renamed, so a material simply gets however many snapshots carry its exact
name. This avoids silently misaligning one material's series onto another.
"""
from __future__ import annotations
import openpyxl

HISTORY_FILE = "/Users/deep/Downloads/Feed Raw Material Price Dashboard..xlsx"

# dated sheet -> canonical snapshot date (YYYY-MM-DD)
SNAPSHOTS = {
    "20260617": "2026-06-17",
    "20260702": "2026-07-02",
    "20260713": "2026-07-13",
    "Copy of 20260722": "2026-07-22",
    "20260802": "2026-08-02",
    "20260824": "2026-08-24",
    "20260902": "2026-09-02",
}


def _num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def _header_cols(rows):
    """Locate the header row and return column indices for product + current price."""
    for i in range(0, min(25, len(rows))):
        row = [str(x) if x is not None else "" for x in rows[i]]
        if any("roduct" in c for c in row):
            cols = {}
            for j, c in enumerate(row):
                cs = c.strip().lower()
                if cs in ("hs", "hs code"):
                    cols["hs"] = j
                elif "product" in cs:
                    cols["product"] = j
                elif "current" in cs and "price" in cs:
                    cols["current"] = j
            return i, cols
    return None, {}


def ingest_price_history() -> dict[str, list[tuple[str, float]]]:
    """Return {NAME: [(date, price), ...]} sorted chronologically.

    NAME is the uppercase/stripped product name. Only snapshots that carry an
    exact name match contribute; missing snapshots are simply absent.

    The history layer is additive/optional — if the snapshot workbook is absent
    or unreadable it returns {} so the core pipeline still runs (honest: a missing
    series is reported as None, never fabricated).
    """
    try:
        wb = openpyxl.load_workbook(HISTORY_FILE, data_only=True)
    except Exception:
        return {}
    out: dict[str, dict[str, float]] = {}
    for sheet, date in SNAPSHOTS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hdr, cols = _header_cols(rows)
        if not cols:
            continue
        p = cols["product"]
        c = cols["current"]
        for r in rows[hdr + 1:]:
            if len(r) <= max(p, c):
                continue
            name = r[p]
            if not name or not str(name).strip():
                continue
            val = _num(r[c])
            if val is None:
                continue
            key = str(name).strip().upper()
            out.setdefault(key, {})[date] = val
    return {k: sorted(v.items()) for k, v in out.items()}
