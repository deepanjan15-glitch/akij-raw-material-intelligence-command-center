"""Load the 87-material master list (Input_Sheet) -> Material entities + category map."""
from __future__ import annotations
import openpyxl
from ..core.entities import Material

MASTER_FILE = "/Users/deep/Desktop/Akij Agro Feed Ltd. Raw Material Price Dashboard..xlsx"

CATEGORY = {
    "MAIZE/CORN": "Grains & Cereals", "CORN FOB": "Grains & Cereals", "CORN CIF": "Grains & Cereals",
    "CORN BASIS": "Grains & Cereals", "CORN FAS": "Grains & Cereals", "CORN CPT": "Grains & Cereals",
    "Corn DDG CIF": "Grains & Cereals", "Corn DDG FOB": "Grains & Cereals",
    "WHEAT BRAN": "Grains & Cereals", "DDGS": "Grains & Cereals",
    "CORN GLUTEN FEED": "Grains & Cereals", "CORN GLUTEN MEAL": "Grains & Cereals",
    "WHEAT CBOT": "Grains & Cereals", "WHEAT 10%": "Grains & Cereals",
    "WHEAT CBOT 10.5%": "Grains & Cereals", "WHEAT CBOT 11.5%": "Grains & Cereals",
    "WHEAT CBOT 12.5%": "Grains & Cereals", "WHEAT AND WHEAT BRAN CBOT 13.5%": "Grains & Cereals",
    "WHEAT CBOT 14.5%": "Grains & Cereals", "WHEAT FEED": "Grains & Cereals",
    "WHEAT FOB 10%": "Grains & Cereals", "WHEAT FOB 10.5%": "Grains & Cereals",
    "WHEAT FOB 11%": "Grains & Cereals", "WHEAT FOB 11.5%": "Grains & Cereals",
    "WHEAT FOB 12.5%": "Grains & Cereals", "WHEAT FOB 13.5%": "Grains & Cereals",
    "WHEAT FOB 14.5%": "Grains & Cereals", "WHEAT FOB 9.5%": "Grains & Cereals",
    "WHEAT CIF 10.5%": "Grains & Cereals", "WHEAT CIF 11.5%": "Grains & Cereals",
    "WHEAT CIF 12.5%": "Grains & Cereals", "WHEAT CIF 13.5%": "Grains & Cereals",
    "WHEAT CIF 14.5%": "Grains & Cereals", "WHEAT CIF Feed": "Grains & Cereals",
    "WHEAT FOB Midds": "Grains & Cereals", "BERLEY": "Grains & Cereals",
    "TAPIOCA RESIDUE PELLETS": "Grains & Cereals",
    "SOYMEAL CIF": "Oilseeds & Meals", "SOYMEAL CIF Hi-Pro": "Oilseeds & Meals",
    "SOYMEAL CIF SMP": "Oilseeds & Meals", "SOYMEAL FOB Hi-Pro": "Oilseeds & Meals",
    "SOYMEAL FOB": "Oilseeds & Meals", "SOYMEAL FOB SMP": "Oilseeds & Meals",
    "SOYMEAL CBOT": "Oilseeds & Meals", "SOYMEAL BASIS": "Oilseeds & Meals",
    "SOYBEAN ORGANIC FEED": "Oilseeds & Meals", "SOYBEAN BASIS": "Oilseeds & Meals",
    "Soybean CBOT": "Oilseeds & Meals", "Soybean CFR": "Oilseeds & Meals",
    "Soybean CIF": "Oilseeds & Meals", "SOYBEAN FOB": "Oilseeds & Meals",
    "Soybean FAS": "Oilseeds & Meals", "Soybean Feed": "Oilseeds & Meals",
    "RAPESEED Meal (RSM)": "Oilseeds & Meals", "RAPESEED / CANOLA CPT": "Oilseeds & Meals",
    "RAPESEED / CANOLA FOB": "Oilseeds & Meals", "RAPESEED EXTRACT": "Oilseeds & Meals",
    "SUNFLOWER MEAL": "Oilseeds & Meals", "SOYBEAN FLOUR": "Oilseeds & Meals",
    "SOYBEAN OIL CFR Crude": "Oils & Fats", "SOYBEAN OIL FOB Crude": "Oils & Fats",
    "SOYBEAN OIL FOB Refined": "Oils & Fats", "RAPESEED OIL FOB Refined": "Oils & Fats",
    "COTTON SEED OIL": "Oils & Fats", "Fish Oil": "Oils & Fats",
    "ANIMAL FEED": "Compound Feed",
    "MONO CALCIUM PHOSPHATE (MCP)": "Minerals", "DICALCIUM PHOSPHATE (DCP)": "Minerals",
    "LIMESTONE POWDER": "Minerals",
    "L-METHIONINE": "Amino Acids", "PH RAW MATERIALS": "Amino Acids",
    "L-THREONINE": "Amino Acids", "L-LYSINE": "Amino Acids", "DL-METHIONINE": "Amino Acids",
    "AMINO ACID (FEED)": "Amino Acids",
    "CHOLINE CHLORIDE": "Additives", "FEED ADDITIVE": "Additives", "FEED PREMIX": "Additives",
    "FEED SUPPLEMENT (LIQUID)": "Protein Meals & Supplements",
    "FEED SUPPLEMENT (POWDER)": "Protein Meals & Supplements",
    "FEED SUPPLEMENT (OTHER)": "Protein Meals & Supplements",
    "FISH MEAL": "Protein Meals & Supplements", "OTHER PROTEIN / MEAL": "Protein Meals & Supplements",
    "POULTRY MEAL": "Protein Meals & Supplements", "SHRIMP MEAL": "Protein Meals & Supplements",
}


def _txt(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ("", "-", "—") else s


def _num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def ingest_materials() -> list[Material]:
    wb = openpyxl.load_workbook(MASTER_FILE, data_only=True)
    ws = wb["Input_Sheet"]
    out: list[Material] = []
    for r in range(13, 100):
        name = _txt(ws.cell(row=r, column=2).value)
        if not name:
            continue
        hs = ws.cell(row=r, column=1).value
        hs = str(int(hs)) if isinstance(hs, float) and hs.is_integer() else (str(hs) if hs else "")
        unit = _txt(ws.cell(row=r, column=3).value) or "MT"
        src = _txt(ws.cell(row=r, column=4).value) or "Other"
        out.append(Material(
            id=f"MAT-{len(out)+1:03d}",
            name=name,
            category=CATEGORY.get(name, "Other"),
            hsCode=hs,
            unit=unit,
            sources=[src],
        ))
    return out


def ingest_benchmarks() -> dict:
    """Return master benchmark snapshot per material name (recomputed, not trusted)."""
    wb = openpyxl.load_workbook(MASTER_FILE, data_only=True)
    ws = wb["Input_Sheet"]
    out = {}
    for r in range(13, 100):
        name = _txt(ws.cell(row=r, column=2).value)
        if not name:
            continue
        out[name] = {
            "lastWeek": _num(ws.cell(row=r, column=5).value),
            "current": _num(ws.cell(row=r, column=6).value),
            "lastMonth": _num(ws.cell(row=r, column=7).value),
            "sixMo": _num(ws.cell(row=r, column=8).value),
            "avg2024": _num(ws.cell(row=r, column=9).value),
            "avg2025": _num(ws.cell(row=r, column=10).value),
            "lastYearYtd": _num(ws.cell(row=r, column=11).value),
            "ytd2026": _num(ws.cell(row=r, column=12).value),
            "trend": _txt(ws.cell(row=r, column=17).value),
            "cheapestCountry1": _txt(ws.cell(row=r, column=20).value),
            "price1": _num(ws.cell(row=r, column=21).value),
            "cheapestCountry2": _txt(ws.cell(row=r, column=22).value),
            "price2": _num(ws.cell(row=r, column=23).value),
            "cheapestCountry3": _txt(ws.cell(row=r, column=24).value),
            "price3": _num(ws.cell(row=r, column=25).value),
            "akijSourcingCountry": _txt(ws.cell(row=r, column=27).value),
            "procurementAction": _txt(ws.cell(row=r, column=29).value),
        }
    return out
