"""Ingest NBR customs import sheets -> normalized Import entities."""
from __future__ import annotations
import openpyxl
from datetime import datetime, date
from ..core.entities import Import, Supplier, country_full

NBR_FILE = "/Users/deep/Desktop/NBR_Raw_Material_Data.xlsx"

# column indices (1-based) in the NBR data header (row 6)
COL = {
    "hs": 10, "desc": 12, "invoice": 13, "net_kg": 16, "qty": 17,
    "uprice": 18, "cur": 21, "coo": 22, "boe": 8, "receipt": 56,
    "exporter": 63, "importer": 5,
}

SHEET_MATERIAL = {
    "CHOLINE CHLORIDE": "Choline Chloride",
    "Corn_Or_Maize": "Maize/Corn",
    "Corn_Gluten Feed_Or CGF": "Corn Gluten Feed/Meal",
    "DCP_MCP_PH_Raw_Mat": "Phosphate (MCP/DCP)",
    "DCP": "Dicalcium Phosphate (DCP)",
    "DDGS": "DDGS",
    "DL-METHIONINE": "Methionine (L/DL)",
    "Feed_Premix_Supliment": "Feed Premix / Supplement",
    "L-Lysine": "L-Lysine",
    "L-THREONINE": "L-Threonine",
    "Limestone_Powder": "Limestone Powder",
    "Rapeseed_Extraction_Or_RSM": "Rapeseed Extraction",
    "Oil_and_Soybean_Meal_Or_SBM": "Soybean Meal/Flour",
    "Soybean_Meal": "Soybean Meal",
    "Wheat": "Wheat",
    "Wheat_Bran": "Wheat Bran",
}


def _num(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(",", "")
        try:
            return float(s)
        except ValueError:
            return None
    return None


def ingest_nbr(loaded_at: str) -> list[Import]:
    wb = openpyxl.load_workbook(NBR_FILE, data_only=True)
    out: list[Import] = []
    n = 0
    for sheet, material_hint in SHEET_MATERIAL.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for r in range(7, ws.max_row + 1):
            desc = ws.cell(row=r, column=COL["desc"]).value
            hs = ws.cell(row=r, column=COL["hs"]).value
            if not desc and not hs:
                continue
            coo = ws.cell(row=r, column=COL["coo"]).value
            cur = ws.cell(row=r, column=COL["cur"]).value
            uprice = _num(ws.cell(row=r, column=COL["uprice"]).value)
            net_kg = _num(ws.cell(row=r, column=COL["net_kg"]).value)
            invoice = _num(ws.cell(row=r, column=COL["invoice"]).value)
            boe = ws.cell(row=r, column=COL["boe"]).value
            receipt = ws.cell(row=r, column=COL["receipt"]).value
            d = receipt if isinstance(receipt, (datetime, date)) else (boe if isinstance(boe, (datetime, date)) else None)
            period = f"{d.year:04d}-{d.month:02d}" if d else ""
            # unit value USD/MT: only for USD-denominated, kg-based unit prices
            unit_value_usd_mt = None
            if cur == "USD" and uprice is not None:
                # Declared U.Price is USD per kg for solid feed items
                unit_value_usd_mt = round(uprice * 1000, 2)
            volume_mt = round(net_kg / 1000, 3) if net_kg else None
            value_usd = round(invoice, 2) if (cur == "USD" and invoice is not None) else None
            n += 1
            out.append(Import(
                id=f"NBR-{sheet[:8]}-{n:04d}",
                materialId=None,  # mapped later
                hsCode=str(hs) if hs else "",
                countryCode=str(coo) if coo else "",
                countryName=country_full(str(coo)) if coo else None,
                description=str(desc).strip(),
                period=period,
                volumeMt=volume_mt,
                valueUsd=value_usd,
                unitValueUsdMt=unit_value_usd_mt,
                exporter=str(ws.cell(row=r, column=COL["exporter"]).value).strip() if ws.cell(row=r, column=COL["exporter"]).value else None,
                importer=str(ws.cell(row=r, column=COL["importer"]).value).strip() if ws.cell(row=r, column=COL["importer"]).value else None,
                loadedAt=loaded_at,
            ))
            out[-1]._sheet = sheet
            out[-1]._material_hint = material_hint
    return out


def ingest_suppliers(imports: list[Import], loaded_at: str) -> list[Supplier]:
    """Aggregate NBR exporters (col 63) into supplier intelligence."""
    from collections import defaultdict
    agg = defaultdict(lambda: {"vol": 0.0, "val": 0.0, "ship": 0, "coo": "", "mats": set()})
    for im in imports:
        if not im.exporter:
            continue
        key = im.exporter.upper()
        a = agg[key]
        a["vol"] += im.volumeMt or 0
        a["val"] += im.valueUsd or 0
        a["ship"] += 1
        a["coo"] = im.countryCode or a["coo"]
        if im.materialId:
            a["mats"].add(im.materialId)
    out = []
    for i, (name, a) in enumerate(sorted(agg.items(), key=lambda kv: kv[1]["vol"], reverse=True)):
        vol = a["vol"]
        out.append(Supplier(
            id=f"SUP-{i+1:04d}",
            name=name.title(),
            countryCode=a["coo"],
            countryName=country_full(a["coo"]),
            materialId=next(iter(a["mats"]), None),
            volumeMt=round(vol, 1),
            valueUsd=round(a["val"], 0),
            unitValueUsdMt=round(a["val"] / vol, 2) if vol else None,
            shipments=a["ship"],
        ))
    return out
